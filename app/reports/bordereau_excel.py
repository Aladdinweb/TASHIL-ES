# COPYRIGHT ILINE TECH 2026 BY FERAK ALADDIN
"""
Générateur Bordereau d'envoi Excel — EPSP ES-SENIA
Template officiel : N° | DESIGNATION DES PIECES | NB RE | OBS
Pied : LE MEDECIN COORDINATEUR (droite) | Pour toutes fins utiles (gauche)
"""
import os
import sys
import datetime
import openpyxl
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill)
from openpyxl.utils import get_column_letter


# ── Couleurs institutionnelles ────────────────────────────────────
BLEU_HEADER  = "1E3A5F"
BLEU_CLAIR   = "D6E4F0"
GRIS_LIGNE   = "F5F5F5"
BLANC        = "FFFFFF"
ORANGE_CAT   = "F0A500"
VERT_TITRE   = "1A5276"


def _bordure(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def _bordure_ext():
    ep = Side(style="medium")
    fi = Side(style="thin")
    return Border(left=ep, right=ep, top=ep, bottom=fi)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _police(bold=False, size=10, color="000000", italic=False):
    return Font(name="Times New Roman", bold=bold,
                size=size, color=color, italic=italic)


def _align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v,
                     wrap_text=wrap)


def generer_bordereau(data: dict, chemin_sortie: str = None) -> str:
    """
    Génère le Bordereau d'envoi en .xlsx.

    data doit contenir :
        numero       : str  — numéro du bordereau (ex: "001/2025")
        date         : str  — date d'émission (JJ/MM/AAAA)
        expediteur   : str  — service émetteur
        destinataire : str  — destinataire
        objet        : str  — objet du bordereau
        lignes       : list de dicts :
            {
              categorie    : "CONGE_ANNUEL" | "CERTIFICAT_MEDICAL_ARRET" | …
              designation  : str  — texte de la pièce
              nb_pieces    : int
              observation  : str  (optionnel)
            }

    Retourne le chemin du fichier généré.
    """
    # ── Chemin de sortie ─────────────────────────────────────────
    if chemin_sortie is None:
        if getattr(sys, 'frozen', False):
            base = os.path.dirname(sys.executable)
        else:
            base = os.path.dirname(
                os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        dossier = os.path.join(base, "bordereaux")
        os.makedirs(dossier, exist_ok=True)
        nom_fichier = (f"Bordereau_{data.get('numero','').replace('/', '-')}"
                       f"_{datetime.date.today().isoformat()}.xlsx")
        chemin_sortie = os.path.join(dossier, nom_fichier)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bordereau d'envoi"

    # ── Largeurs colonnes ─────────────────────────────────────────
    # A=N°  B=Désignation  C=NB RE  D=OBS
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 68
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 24

    ligne = 1

    # ════════════════════════════════════════════════════════════
    # EN-TÊTE INSTITUTIONNEL
    # ════════════════════════════════════════════════════════════

    # Ligne 1 : République / Ministère (sur 2 colonnes)
    ws.merge_cells(f"A{ligne}:D{ligne}")
    c = ws[f"A{ligne}"]
    c.value = ("REPUBLIQUE ALGERIENNE DEMOCRATIQUE ET POPULAIRE\n"
               "MINISTERE DE LA SANTE")
    c.font      = _police(bold=True, size=10)
    c.alignment = _align(h="center")
    c.fill      = _fill(BLEU_HEADER)
    c.font      = Font(name="Times New Roman", bold=True,
                       size=10, color=BLANC)
    ws.row_dimensions[ligne].height = 28
    ligne += 1

    # Ligne 2 : EPSP
    ws.merge_cells(f"A{ligne}:D{ligne}")
    c = ws[f"A{ligne}"]
    c.value     = "ETABLISSEMENT PUBLIC DE SANTE DE PROXIMITE"
    c.font      = Font(name="Times New Roman", bold=True,
                       size=11, color=BLANC)
    c.alignment = _align(h="center")
    c.fill      = _fill(BLEU_HEADER)
    ws.row_dimensions[ligne].height = 20
    ligne += 1

    # Ligne 3 : Nom EPSP
    ws.merge_cells(f"A{ligne}:D{ligne}")
    c = ws[f"A{ligne}"]
    c.value     = "ES-SENIA"
    c.font      = Font(name="Times New Roman", bold=True,
                       size=14, color=BLANC)
    c.alignment = _align(h="center")
    c.fill      = _fill(BLEU_HEADER)
    ws.row_dimensions[ligne].height = 26
    ligne += 1

    # Ligne 4 : séparateur
    ws.merge_cells(f"A{ligne}:D{ligne}")
    ws[f"A{ligne}"].fill = _fill(ORANGE_CAT)
    ws.row_dimensions[ligne].height = 4
    ligne += 1

    # Ligne 5 : infos expéditeur / date
    ws[f"A{ligne}"].value = "DE :"
    ws[f"A{ligne}"].font  = _police(bold=True, size=9)
    ws.merge_cells(f"B{ligne}:C{ligne}")
    ws[f"B{ligne}"].value     = data.get("expediteur", "")
    ws[f"B{ligne}"].font      = _police(bold=True, size=10)
    ws[f"B{ligne}"].alignment = _align(h="left")
    ws[f"D{ligne}"].value     = f"Date : {data.get('date', '')}"
    ws[f"D{ligne}"].font      = _police(bold=True, size=9)
    ws[f"D{ligne}"].alignment = _align(h="right")
    ws.row_dimensions[ligne].height = 18
    ligne += 1

    # Ligne 6 : Destinataire
    ws[f"A{ligne}"].value = "A :"
    ws[f"A{ligne}"].font  = _police(bold=True, size=9)
    ws.merge_cells(f"B{ligne}:D{ligne}")
    ws[f"B{ligne}"].value     = data.get("destinataire", "Direction")
    ws[f"B{ligne}"].font      = _police(bold=True, size=10)
    ws[f"B{ligne}"].alignment = _align(h="left")
    ws.row_dimensions[ligne].height = 18
    ligne += 1

    # Ligne 7 : Numéro
    ws[f"A{ligne}"].value = "N° :"
    ws[f"A{ligne}"].font  = _police(bold=True, size=9)
    ws.merge_cells(f"B{ligne}:C{ligne}")
    ws[f"B{ligne}"].value = data.get("numero", "")
    ws[f"B{ligne}"].font  = _police(bold=True, size=10)
    ws[f"B{ligne}"].alignment = _align(h="left")
    ws.row_dimensions[ligne].height = 18
    ligne += 1

    # Ligne 8 : Objet
    ws[f"A{ligne}"].value = "OBJET :"
    ws[f"A{ligne}"].font  = _police(bold=True, size=9)
    ws.merge_cells(f"B{ligne}:D{ligne}")
    ws[f"B{ligne}"].value     = data.get("objet",
                                         "TRANSMISSION DE DOCUMENTS")
    ws[f"B{ligne}"].font      = _police(bold=True, size=10)
    ws[f"B{ligne}"].alignment = _align(h="left")
    ws.row_dimensions[ligne].height = 20
    ligne += 1

    # ════════════════════════════════════════════════════════════
    # TITRE DU TABLEAU
    # ════════════════════════════════════════════════════════════
    ws.merge_cells(f"A{ligne}:D{ligne}")
    c = ws[f"A{ligne}"]
    c.value     = "BORDEREAU D'ENVOI"
    c.font      = Font(name="Times New Roman", bold=True,
                       size=13, color=BLANC)
    c.alignment = _align(h="center")
    c.fill      = _fill(VERT_TITRE)
    c.border    = _bordure("medium")
    ws.row_dimensions[ligne].height = 26
    ligne += 1

    # ════════════════════════════════════════════════════════════
    # EN-TÊTES COLONNES
    # ════════════════════════════════════════════════════════════
    entetes = ["N°", "DESIGNATION DES PIECES", "NB\nRE", "OBS"]
    for col_idx, entete in enumerate(entetes, start=1):
        cell = ws.cell(row=ligne, column=col_idx)
        cell.value     = entete
        cell.font      = Font(name="Times New Roman", bold=True,
                              size=10, color=BLANC)
        cell.alignment = _align(h="center")
        cell.fill      = _fill(BLEU_HEADER)
        cell.border    = _bordure("medium")
    ws.row_dimensions[ligne].height = 28
    ligne += 1

    # ════════════════════════════════════════════════════════════
    # CORPS — Lignes par catégorie
    # ════════════════════════════════════════════════════════════
    CATEGORIES_ORDRE = [
        ("CONGE_ANNUEL",              "CONGE ANNUEL"),
        ("CERTIFICAT_MEDICAL_ARRET",  "CERTIFICAT MEDICAL D'ARRET DE TRAVAIL"),
        ("CERTIFICAT_MEDICAL_REPRISE","CERTIFICAT MEDICAL DE REPRISE"),
        ("DEMANDE_3_JOURS_NAISSANCE", "DEMANDE DE 3 JOURS DE NAISSANCE"),
        ("DEMANDE_ANNULATION_CONGE",  "DEMANDE D'ANNULATION DE CONGE"),
    ]

    lignes_data = data.get("lignes", [])
    numero_piece = 1
    ligne_depart_tableau = ligne

    for cat_cle, cat_label in CATEGORIES_ORDRE:
        pieces_cat = [l for l in lignes_data
                      if l.get("categorie") == cat_cle]
        if not pieces_cat:
            continue

        # ── Ligne de catégorie ────────────────────────────────
        ws.merge_cells(f"A{ligne}:D{ligne}")
        c = ws[f"A{ligne}"]
        c.value     = f"  {cat_label}"
        c.font      = Font(name="Times New Roman", bold=True,
                           size=10, color="FFFFFF")
        c.alignment = _align(h="left")
        c.fill      = _fill("2E4057")
        c.border    = _bordure("thin")
        ws.row_dimensions[ligne].height = 20
        ligne += 1

        # ── Pièces de la catégorie ────────────────────────────
        for idx, piece in enumerate(pieces_cat):
            bg = BLANC if idx % 2 == 0 else GRIS_LIGNE

            # N°
            c_num = ws.cell(row=ligne, column=1)
            c_num.value     = str(numero_piece)
            c_num.font      = _police(bold=True, size=10)
            c_num.alignment = _align(h="center")
            c_num.fill      = _fill(bg)
            c_num.border    = _bordure("thin")

            # Désignation
            c_des = ws.cell(row=ligne, column=2)
            c_des.value     = piece.get("designation", "")
            c_des.font      = _police(size=10)
            c_des.alignment = _align(h="left", wrap=True)
            c_des.fill      = _fill(bg)
            c_des.border    = _bordure("thin")

            # NB RE
            c_nb = ws.cell(row=ligne, column=3)
            c_nb.value     = piece.get("nb_pieces", 1)
            c_nb.font      = _police(bold=True, size=10)
            c_nb.alignment = _align(h="center")
            c_nb.fill      = _fill(bg)
            c_nb.border    = _bordure("thin")

            # OBS
            c_obs = ws.cell(row=ligne, column=4)
            c_obs.value     = piece.get("observation",
                                        "Pour toutes fins utiles")
            c_obs.font      = _police(size=9, italic=True)
            c_obs.alignment = _align(h="center", wrap=True)
            c_obs.fill      = _fill(bg)
            c_obs.border    = _bordure("thin")

            # Hauteur selon longueur désignation
            longueur = len(str(piece.get("designation", "")))
            hauteur  = max(30, min(80, (longueur // 60 + 1) * 18))
            ws.row_dimensions[ligne].height = hauteur

            numero_piece += 1
            ligne += 1

    # ════════════════════════════════════════════════════════════
    # PIED DE PAGE
    # ════════════════════════════════════════════════════════════
    # Ligne vide
    ws.merge_cells(f"A{ligne}:D{ligne}")
    ws.row_dimensions[ligne].height = 10
    ligne += 1

    # Séparateur
    ws.merge_cells(f"A{ligne}:D{ligne}")
    ws[f"A{ligne}"].fill   = _fill(BLEU_HEADER)
    ws.row_dimensions[ligne].height = 3
    ligne += 1

    # Pied : OBS gauche | Médecin coordinateur droite
    ws.merge_cells(f"A{ligne}:B{ligne}")
    c_obs_pied = ws[f"A{ligne}"]
    c_obs_pied.value     = "OBS : « Pour toutes fins utiles »"
    c_obs_pied.font      = Font(name="Times New Roman", bold=True,
                                size=10, italic=True)
    c_obs_pied.alignment = _align(h="left", v="top")
    c_obs_pied.border    = _bordure("thin")
    ws.row_dimensions[ligne].height = 60

    ws.merge_cells(f"C{ligne}:D{ligne}")
    c_med = ws[f"C{ligne}"]
    c_med.value     = "LE MEDECIN COORDINATEUR"
    c_med.font      = Font(name="Times New Roman", bold=True, size=11)
    c_med.alignment = Alignment(horizontal="center", vertical="top",
                                wrap_text=True)
    c_med.border    = _bordure("thin")
    ligne += 1

    # Espace signature médecin
    ws.merge_cells(f"C{ligne}:D{ligne}")
    c_sig = ws[f"C{ligne}"]
    c_sig.value     = "\n\n\n"
    c_sig.border    = _bordure("thin")
    ws.row_dimensions[ligne].height = 55
    ligne += 1

    # Ligne finale
    ws.merge_cells(f"A{ligne}:D{ligne}")
    ws[f"A{ligne}"].fill = _fill(BLEU_HEADER)
    ws.row_dimensions[ligne].height = 4

    # ── Figer les en-têtes ────────────────────────────────────
    ws.freeze_panes = f"A{ligne_depart_tableau}"

    # ── Zone d'impression ─────────────────────────────────────
    ws.print_area = f"A1:D{ligne}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_margins.left      = 0.5
    ws.page_margins.right     = 0.5
    ws.page_margins.top       = 0.75
    ws.page_margins.bottom    = 0.75

    wb.save(chemin_sortie)
    return chemin_sortie
